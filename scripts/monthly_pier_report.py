import sqlite3
import pandas as pd
import base64
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl.styles import Font, Alignment
from agentmail import AgentMail

def ticks_to_datetime(ticks):
    return datetime(1, 1, 1) + timedelta(microseconds=ticks // 10)

def get_prev_month_range():
    today = datetime.today()
    first_day_this_month = datetime(today.year, today.month, 1)
    last_month = first_day_this_month - timedelta(days=1)
    first_day_last_month = datetime(last_month.year, last_month.month, 1)
    
    # .NET ticks calculation
    start_tick = int(first_day_last_month.timestamp() * 10000000) + 621355968000000000
    end_tick = int(first_day_this_month.timestamp() * 10000000) + 621355968000000000
    return start_tick, end_tick, first_day_last_month.strftime("%B %Y")

def main():
    WORKSPACE = Path("/Users/ryantaylorvegh/.openclaw/workspace")
    db_path = WORKSPACE / "backup_2026-03-13_v4.db"
    excel_path = WORKSPACE / "The_PIER_Report.xlsx"
    
    conn = sqlite3.connect(db_path)
    start_tick, end_tick, month_name = get_prev_month_range()
    
    # Query logic
    query = """
    SELECT 
      WU.Start, 
      WU.End, 
      (WU.Duration / 36000000000.0) as DurationHours, 
      WU.Description, 
      GROUP_CONCAT(T.Value, ', ') as Tags 
    FROM WorkUnits WU 
    JOIN Projects P ON WU.ProjectId = P.ProjectId
    LEFT JOIN WorkUnitTags WUT ON WU.WorkUnitId = WUT.WorkUnitId 
    LEFT JOIN Tags T ON WUT.TagId = T.TagId 
    WHERE P.Name = 'The PIER'
      AND WU.Start >= ? 
      AND WU.Start < ?
    GROUP BY WU.WorkUnitId
    """
    df = pd.read_sql_query(query, conn, params=(start_tick, end_tick))
    
    # Transform
    df['Start'] = df['Start'].apply(lambda x: ticks_to_datetime(x).strftime("%Y-%m-%d %H:%M"))
    df['End'] = df['End'].apply(lambda x: ticks_to_datetime(x).strftime("%Y-%m-%d %H:%M"))
    
    # Total
    total_hours = df['DurationHours'].sum()
    df = df[['Start', 'End', 'DurationHours', 'Tags', 'Description']]
    
    # Excel Generation
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Hours')
        wb = writer.book
        ws = writer.sheets['Hours']
        
        # Add Total Row
        new_row = len(df) + 2
        ws.cell(row=new_row, column=2, value="Total Hours:")
        ws.cell(row=new_row, column=3, value=f"{total_hours:.2f}")
        
        # Formatting
        font = Font(name='Arial', size=11)
        for row in ws.iter_rows():
            for cell in row:
                cell.font = font
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Bold Header + Total
        for cell in ws[1]:
            cell.font = Font(name='Arial', size=11, bold=True)
        for cell in ws[new_row]:
            cell.font = Font(name='Arial', size=11, bold=True)
            
        for col in ws.columns:
            column = col[0].column_letter
            ws.column_dimensions[column].width = min(max(20, 20), 50)

    # Email
    env = {}
    with open(WORKSPACE / ".secrets/agentmail.env") as f:
        for line in f:
            if '=' in line:
                k, v = line.strip().split('=', 1)
                env[k] = v.strip('"\'')
    
    client = AgentMail(api_key=env.get("AGENTMAIL_API_KEY"))
    client.inboxes.messages.send(
        inbox_id=env.get("AGENTMAIL_FROM_INBOX") or "sallysquirrel@agentmail.to",
        to="djholtrop@gmail.com",
        cc=["ryan.vegh@gmail.com"],
        subject=f"The PIER Report - {month_name}",
        text=f"Hi Dave, attached is the monthly PIER report for {month_name}.",
        attachments=[{"filename": f"PIER_Report_{month_name.replace(' ', '_')}.xlsx", "content": base64.b64encode(excel_path.read_bytes()).decode()}]
    )

if __name__ == "__main__":
    main()
