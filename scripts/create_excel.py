import pandas as pd
from datetime import datetime
from pathlib import Path
from openpyxl.styles import Font, Alignment

WORKSPACE = Path("/Users/ryantaylorvegh/.openclaw/workspace")
csv_file = WORKSPACE / "2026 02 February Hours with Total.csv"
excel_file = WORKSPACE / "2026 02 February Hours.xlsx"

# Load the CSV
df = pd.read_csv(csv_file)

# Reorder columns: Start, End, DurationHours, Tags, Description
df = df[['Start', 'End', 'DurationHours', 'Tags', 'Description']]

# Write to Excel
with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name='February Hours')
    
    workbook = writer.book
    worksheet = writer.sheets['February Hours']
    
    # 1. Font consistency
    font = Font(name='Arial', size=11)
    for row in worksheet.iter_rows():
        for cell in row:
            cell.font = font
            
    # 2. Bold headers
    header_font = Font(name='Arial', size=11, bold=True)
    for cell in worksheet[1]:
        cell.font = header_font
        
    # 3. Bold Total row
    for row in worksheet.iter_rows(min_row=2):
        if row[2].value and "Total Hours" in str(row[2].value):
            for cell in row:
                cell.font = Font(name='Arial', size=11, bold=True)
                
    # 4. Autofit column widths & Wrap text
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        # Generous width calculation
        adjusted_width = min(max(max_length + 5, 15), 50)
        worksheet.column_dimensions[column].width = adjusted_width

    # 5. Format Date/Time columns with AM/PM (using cell.number_format)
    # Pandas wrote these as text, so we may need to convert them to real datetime objects first for Excel to format them
    for row in worksheet.iter_rows(min_row=2):
        for cell in [row[0], row[1]]:
            if cell.value:
                try:
                    cell.value = datetime.strptime(str(cell.value), '%Y-%m-%d %H:%M')
                    cell.number_format = 'yyyy-mm-dd hh:mm AM/PM'
                except:
                    pass



print(f"Excel created at {excel_file}")
